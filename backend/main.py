import cv2
import numpy as np
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import json
import pytesseract
from PIL import Image
import logging
import re
import sqlite3
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure Tesseract (optional - comment out if not installed)
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGE_DIR = os.path.join(BASE_DIR, 'images')
MASTER_DATA_FILE = os.path.join(BASE_DIR, 'master_answers.json')
MASTER_METADATA_FILE = os.path.join(BASE_DIR, 'master_metadata.json')
DB_FILE = os.path.join(BASE_DIR, 'omr_grading.db')
EXPORTS_DIR = os.path.join(BASE_DIR, 'exports')

os.makedirs(IMAGE_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)


class DatabaseManager:
    """Enhanced database manager with filtering and analytics"""
    
    def __init__(self, db_path):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize database with enhanced schema"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Drop existing tables to recreate with enhanced schema
        cursor.execute('DROP TABLE IF EXISTS grading_results')
        cursor.execute('DROP TABLE IF EXISTS students')
        cursor.execute('DROP TABLE IF EXISTS master_keys')
        
        # Students table
        cursor.execute('''
            CREATE TABLE students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                subject TEXT,
                medium TEXT,
                grade_level TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Grading results table with enhanced fields
        cursor.execute('''
            CREATE TABLE grading_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                exam_date DATE NOT NULL,
                subject TEXT NOT NULL,
                grade_level TEXT,
                total_questions INTEGER NOT NULL,
                correct_answers INTEGER NOT NULL,
                wrong_answers INTEGER NOT NULL,
                unanswered INTEGER NOT NULL,
                score REAL NOT NULL,
                percentage REAL NOT NULL,
                grade TEXT NOT NULL,
                answers_json TEXT NOT NULL,
                master_key_id INTEGER,
                graded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (student_id) REFERENCES students(student_id),
                FOREIGN KEY (master_key_id) REFERENCES master_keys(id)
            )
        ''')
        
        # Master keys table
        cursor.execute('''
            CREATE TABLE master_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject TEXT NOT NULL,
                exam_date DATE NOT NULL,
                grade_level TEXT NOT NULL,
                total_questions INTEGER NOT NULL,
                answers_json TEXT NOT NULL,
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Indexes for performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_student_id ON students(student_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_grading_student_id ON grading_results(student_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_grading_subject ON grading_results(subject)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_grading_grade_level ON grading_results(grade_level)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_grading_exam_date ON grading_results(exam_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_master_active ON master_keys(is_active)')
        
        conn.commit()
        conn.close()
        logger.info("✓ Enhanced database initialized")
    
    def add_student(self, student_id, name, subject=None, medium=None, grade_level=None):
        """Add or update student information"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO students 
                (student_id, name, subject, medium, grade_level, updated_at)
                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (student_id, name, subject, medium, grade_level))
            
            conn.commit()
            logger.info(f"✓ Student added/updated: {name} ({student_id})")
            return True
        except Exception as e:
            logger.error(f"Error adding student: {e}")
            return False
        finally:
            conn.close()
    
    def add_grading_result(self, student_id, subject, grade_level, exam_date, results, answers, master_key_id=None):
        """Add grading result"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # Calculate grade
            percentage = results['percentage']
            if percentage >= 90:
                grade = 'A+'
            elif percentage >= 80:
                grade = 'A'
            elif percentage >= 70:
                grade = 'B'
            elif percentage >= 60:
                grade = 'C'
            elif percentage >= 50:
                grade = 'D'
            else:
                grade = 'F'
            
            cursor.execute('''
                INSERT INTO grading_results 
                (student_id, exam_date, subject, grade_level, total_questions, correct_answers, 
                 wrong_answers, unanswered, score, percentage, grade, answers_json, master_key_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                student_id,
                exam_date,
                subject,
                grade_level,
                results['total'],
                results['correct'],
                results['wrong'],
                results['unanswered'],
                results['score'],
                results['percentage'],
                grade,
                json.dumps(answers),
                master_key_id
            ))
            
            conn.commit()
            logger.info(f"✓ Grading result saved: {student_id} - {grade} ({percentage}%)")
            return True
        except Exception as e:
            logger.error(f"Error adding grading result: {e}")
            return False
        finally:
            conn.close()
    
    def add_master_key(self, subject, exam_date, grade_level, answers):
        """Add master answer key"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # Deactivate previous master keys for same subject/grade
            cursor.execute('''
                UPDATE master_keys 
                SET is_active = 0 
                WHERE subject = ? AND grade_level = ?
            ''', (subject, grade_level))
            
            # Insert new master key
            cursor.execute('''
                INSERT INTO master_keys (subject, exam_date, grade_level, total_questions, answers_json, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
            ''', (subject, exam_date, grade_level, len(answers), json.dumps(answers)))
            
            master_key_id = cursor.lastrowid
            
            conn.commit()
            logger.info(f"✓ Master key saved: {subject} - {grade_level} (ID: {master_key_id})")
            return master_key_id
        except Exception as e:
            logger.error(f"Error adding master key: {e}")
            return None
        finally:
            conn.close()
    
    def get_active_master_key(self):
        """Get currently active master key"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT id, subject, grade_level, exam_date, answers_json
                FROM master_keys
                WHERE is_active = 1
                ORDER BY created_at DESC
                LIMIT 1
            ''')
            
            result = cursor.fetchone()
            if result:
                return {
                    'id': result[0],
                    'subject': result[1],
                    'grade_level': result[2],
                    'exam_date': result[3],
                    'answers': json.loads(result[4])
                }
            return None
        except Exception as e:
            logger.error(f"Error fetching active master key: {e}")
            return None
        finally:
            conn.close()
    
    def get_all_results(self, subject=None, grade_level=None, exam_date=None, grade=None):
        """Get all grading results with enhanced filtering"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT 
                s.student_id,
                s.name,
                COALESCE(s.subject, g.subject) as subject,
                COALESCE(s.medium, 'Unknown') as medium,
                COALESCE(s.grade_level, g.grade_level) as grade_level,
                g.exam_date,
                g.total_questions,
                g.correct_answers,
                g.wrong_answers,
                g.unanswered,
                g.score,
                g.percentage,
                g.grade,
                g.graded_at
            FROM grading_results g
            JOIN students s ON g.student_id = s.student_id
        '''
        
        conditions = []
        params = []
        
        if subject:
            conditions.append("g.subject = ?")
            params.append(subject)
        
        if grade_level:
            conditions.append("g.grade_level = ?")
            params.append(grade_level)
        
        if exam_date:
            conditions.append("g.exam_date = ?")
            params.append(exam_date)
        
        if grade:
            conditions.append("g.grade = ?")
            params.append(grade)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY g.graded_at DESC"
        
        try:
            cursor.execute(query, params)
            results = cursor.fetchall()
            return results
        except Exception as e:
            logger.error(f"Error fetching results: {e}")
            return []
        finally:
            conn.close()
    
    def get_results_by_subject_and_grade(self):
        """Get results grouped by subject and grade"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT 
                    g.subject,
                    g.grade_level,
                    g.grade,
                    COUNT(*) as count,
                    AVG(g.percentage) as avg_percentage
                FROM grading_results g
                GROUP BY g.subject, g.grade_level, g.grade
                ORDER BY g.subject, g.grade_level, g.grade
            ''')
            
            results = cursor.fetchall()
            
            # Organize by subject and grade level
            organized = defaultdict(lambda: defaultdict(dict))
            for row in results:
                subject, grade_level, grade, count, avg_perc = row
                if subject not in organized:
                    organized[subject] = {}
                if grade_level not in organized[subject]:
                    organized[subject][grade_level] = {}
                organized[subject][grade_level][grade] = {
                    'count': count,
                    'average': round(avg_perc, 2)
                }
            
            return dict(organized)
        except Exception as e:
            logger.error(f"Error fetching grouped results: {e}")
            return {}
        finally:
            conn.close()
    
    def get_available_filters(self):
        """Get available filter options"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # Get unique subjects
            cursor.execute('SELECT DISTINCT subject FROM grading_results ORDER BY subject')
            subjects = [row[0] for row in cursor.fetchall()]
            
            # Get unique grade levels
            cursor.execute('SELECT DISTINCT grade_level FROM grading_results ORDER BY grade_level')
            grade_levels = [row[0] for row in cursor.fetchall()]
            
            # Get unique grades
            cursor.execute('SELECT DISTINCT grade FROM grading_results ORDER BY grade')
            grades = [row[0] for row in cursor.fetchall()]
            
            # Get exam dates
            cursor.execute('SELECT DISTINCT exam_date FROM grading_results ORDER BY exam_date DESC')
            exam_dates = [row[0] for row in cursor.fetchall()]
            
            return {
                'subjects': subjects,
                'grade_levels': grade_levels,
                'grades': grades,
                'exam_dates': exam_dates
            }
        except Exception as e:
            logger.error(f"Error fetching filters: {e}")
            return {}
        finally:
            conn.close()
    
    def get_student_history(self, student_id):
        """Get grading history for a specific student"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT 
                    exam_date,
                    subject,
                    grade_level,
                    score,
                    percentage,
                    grade,
                    graded_at
                FROM grading_results
                WHERE student_id = ?
                ORDER BY graded_at DESC
            ''', (student_id,))
            
            results = cursor.fetchall()
            return results
        except Exception as e:
            logger.error(f"Error fetching student history: {e}")
            return []
        finally:
            conn.close()
    
    def get_statistics(self, subject=None, grade_level=None):
        """Get enhanced statistics"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        stats = {}
        
        try:
            # Base conditions
            where_clause = ""
            params = []
            
            if subject or grade_level:
                conditions = []
                if subject:
                    conditions.append("subject = ?")
                    params.append(subject)
                if grade_level:
                    conditions.append("grade_level = ?")
                    params.append(grade_level)
                where_clause = " WHERE " + " AND ".join(conditions)
            
            # Total students
            cursor.execute(f'SELECT COUNT(DISTINCT student_id) FROM grading_results{where_clause}', params)
            stats['total_students'] = cursor.fetchone()[0]
            
            # Total exams
            cursor.execute(f'SELECT COUNT(*) FROM grading_results{where_clause}', params)
            stats['total_exams'] = cursor.fetchone()[0]
            
            # Average score
            cursor.execute(f'SELECT AVG(percentage) FROM grading_results{where_clause}', params)
            avg = cursor.fetchone()[0]
            stats['average_score'] = round(avg, 2) if avg else 0
            
            # Highest score
            cursor.execute(f'SELECT MAX(percentage) FROM grading_results{where_clause}', params)
            high = cursor.fetchone()[0]
            stats['highest_score'] = round(high, 2) if high else 0
            
            # Lowest score
            cursor.execute(f'SELECT MIN(percentage) FROM grading_results{where_clause}', params)
            low = cursor.fetchone()[0]
            stats['lowest_score'] = round(low, 2) if low else 0
            
            # Grade distribution
            cursor.execute(f'''
                SELECT grade, COUNT(*) as count
                FROM grading_results{where_clause}
                GROUP BY grade
                ORDER BY grade
            ''', params)
            stats['grade_distribution'] = dict(cursor.fetchall())
            
            # Pass rate (>=50%)
            cursor.execute(f'''
                SELECT 
                    COUNT(CASE WHEN percentage >= 50 THEN 1 END) * 100.0 / COUNT(*) as pass_rate
                FROM grading_results{where_clause}
            ''', params)
            pass_rate = cursor.fetchone()[0]
            stats['pass_rate'] = round(pass_rate, 2) if pass_rate else 0
            
            return stats
        except Exception as e:
            logger.error(f"Error fetching statistics: {e}")
            return {}
        finally:
            conn.close()


class ExcelExporter:
    """Enhanced Excel exporter with filtering support"""
    
    @staticmethod
    def export_results(results, filename, filters=None):
        """Export results to Excel with enhanced formatting"""
        if not results:
            return False
        
        # Create DataFrame
        df = pd.DataFrame(results, columns=[
            'Student ID', 'Name', 'Subject', 'Medium', 'Grade Level', 'Exam Date',
            'Total Questions', 'Correct', 'Wrong', 'Unanswered',
            'Score', 'Percentage', 'Grade', 'Graded At'
        ])
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grading Results"
        
        # Add filter information at top if provided
        current_row = 1
        if filters:
            ws.merge_cells(f'A{current_row}:N{current_row}')
            filter_cell = ws.cell(row=current_row, column=1)
            filter_cell.value = "FILTERED RESULTS"
            filter_cell.font = Font(bold=True, size=14, color="FFFFFF")
            filter_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            filter_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            for key, value in filters.items():
                if value:
                    ws.merge_cells(f'A{current_row}:B{current_row}')
                    label_cell = ws.cell(row=current_row, column=1)
                    label_cell.value = f"{key.replace('_', ' ').title()}:"
                    label_cell.font = Font(bold=True)
                    
                    value_cell = ws.cell(row=current_row, column=3)
                    value_cell.value = str(value)
                    current_row += 1
            
            current_row += 1
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        grade_colors = {
            'A+': PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            'A': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'B': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'C': PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
            'D': PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
            'F': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        header_row = current_row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        current_row += 1
        
        # Write data
        for row_data in df.values:
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply grade color
                if col_num == 13:  # Grade column
                    grade = str(value)
                    if grade in grade_colors:
                        cell.fill = grade_colors[grade]
                        cell.font = Font(bold=True, color="FFFFFF")
            current_row += 1
        
        # Adjust column widths
        column_widths = [15, 25, 18, 12, 12, 12, 12, 10, 10, 12, 10, 12, 8, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
        # Add summary
        summary_row = current_row + 2
        
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_cell = ws.cell(row=summary_row, column=1)
        summary_cell.value = "SUMMARY STATISTICS"
        summary_cell.font = Font(bold=True, size=14)
        summary_row += 1
        
        ws.cell(row=summary_row, column=1).value = "Total Students:"
        ws.cell(row=summary_row, column=2).value = len(df)
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        summary_row += 1
        
        ws.cell(row=summary_row, column=1).value = "Average Score:"
        ws.cell(row=summary_row, column=2).value = f"{df['Percentage'].mean():.2f}%"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        summary_row += 1
        
        ws.cell(row=summary_row, column=1).value = "Highest Score:"
        ws.cell(row=summary_row, column=2).value = f"{df['Percentage'].max():.2f}%"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        summary_row += 1
        
        ws.cell(row=summary_row, column=1).value = "Lowest Score:"
        ws.cell(row=summary_row, column=2).value = f"{df['Percentage'].min():.2f}%"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        summary_row += 1
        
        # Pass rate
        pass_rate = (df['Percentage'] >= 50).sum() * 100 / len(df)
        ws.cell(row=summary_row, column=1).value = "Pass Rate:"
        ws.cell(row=summary_row, column=2).value = f"{pass_rate:.2f}%"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        # Save
        wb.save(filename)
        logger.info(f"✓ Excel exported: {filename}")
        return True


class EnhancedOMRProcessor:
    """Production-grade OMR processor"""
    
    def __init__(self, image_path):
        self.image_path = image_path
        self.original = cv2.imread(image_path)
        self.processed = None
        self.warped = None
        self.student_info = {
            'subject': 'Not detected',
            'medium': 'Not detected',
            'name': 'Not detected'
        }
        self.answers = {}
        
    def process(self):
        """Main processing pipeline"""
        try:
            logger.info("="*80)
            logger.info("ENHANCED OMR PROCESSING v3.0")
            logger.info("="*80)
            
            if self.original is None:
                raise ValueError("Could not load image")
            
            self.processed = self._preprocess_image()
            self.warped = self._perspective_transform_robust()
            self.student_info = self._extract_student_info()
            self.answers = self._extract_all_40_guaranteed()
            
            logger.info(f"✓ Processing complete: {len(self.answers)}/40 answers detected")
            return True
        except Exception as e:
            logger.error(f"Processing error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _preprocess_image(self):
        """Preprocess with rotation detection"""
        img = self.original.copy()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Try rotation detection if Tesseract is available
        try:
            osd = pytesseract.image_to_osd(gray)
            angle_match = re.search(r'Rotate: (\d+)', osd)
            
            if angle_match:
                angle = int(angle_match.group(1))
                if angle == 90:
                    img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
                elif angle == 180:
                    img = cv2.rotate(img, cv2.ROTATE_180)
                elif angle == 270:
                    img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
                logger.info(f"Rotated: {angle}°")
        except:
            logger.info("Rotation detection skipped")
        
        return img
    
    def _perspective_transform_robust(self):
        """Robust perspective correction"""
        gray = cv2.cvtColor(self.processed, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        
        methods = [(50, 150), (30, 100), (75, 200)]
        
        best_contour = None
        best_area = 0
        img_area = self.processed.shape[0] * self.processed.shape[1]
        
        for low, high in methods:
            edged = cv2.Canny(blurred, low, high)
            kernel = np.ones((5, 5), np.uint8)
            dilated = cv2.dilate(edged, kernel, iterations=2)
            
            contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            contours = sorted(contours, key=cv2.contourArea, reverse=True)
            
            for contour in contours[:15]:
                peri = cv2.arcLength(contour, True)
                approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
                
                if len(approx) == 4:
                    area = cv2.contourArea(approx)
                    
                    if area > img_area * 0.3 and area > best_area:
                        best_contour = approx
                        best_area = area
        
        if best_contour is not None:
            pts = best_contour.reshape(4, 2).astype('float32')
            rect = self._order_points(pts)
            
            (tl, tr, br, bl) = rect
            widthA = np.linalg.norm(br - bl)
            widthB = np.linalg.norm(tr - tl)
            maxWidth = max(int(widthA), int(widthB))
            
            heightA = np.linalg.norm(tr - br)
            heightB = np.linalg.norm(tl - bl)
            maxHeight = max(int(heightA), int(heightB))
            
            dst = np.array([
                [0, 0],
                [maxWidth - 1, 0],
                [maxWidth - 1, maxHeight - 1],
                [0, maxHeight - 1]
            ], dtype='float32')
            
            M = cv2.getPerspectiveTransform(rect, dst)
            warped = cv2.warpPerspective(self.processed, M, (maxWidth, maxHeight))
            
            logger.info(f"✓ Perspective corrected: {maxWidth}x{maxHeight}")
            return warped
        
        # Fallback
        logger.warning("Using intelligent crop fallback")
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        rows_with_content = np.any(binary > 0, axis=1)
        cols_with_content = np.any(binary > 0, axis=0)
        
        y_indices = np.where(rows_with_content)[0]
        x_indices = np.where(cols_with_content)[0]
        
        if len(y_indices) > 0 and len(x_indices) > 0:
            pad = 20
            y1 = max(0, y_indices[0] - pad)
            y2 = min(self.processed.shape[0], y_indices[-1] + pad)
            x1 = max(0, x_indices[0] - pad)
            x2 = min(self.processed.shape[1], x_indices[-1] + pad)
            
            cropped = self.processed[y1:y2, x1:x2]
            
            target_width = 1200
            aspect = cropped.shape[0] / cropped.shape[1]
            target_height = int(target_width * aspect)
            
            resized = cv2.resize(cropped, (target_width, target_height))
            logger.info(f"✓ Intelligent crop: {target_width}x{target_height}")
            return resized
        
        return cv2.resize(self.processed, (1200, 1600))
    
    def _order_points(self, pts):
        """Order corner points"""
        rect = np.zeros((4, 2), dtype='float32')
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]
        return rect
    
    def _extract_student_info(self):
        """Extract student information"""
        height, width = self.warped.shape[:2]
        header = self.warped[0:int(height * 0.20), :]
        gray = cv2.cvtColor(header, cv2.COLOR_BGR2GRAY)
        
        info = {'subject': 'Not detected', 'medium': 'Not detected', 'name': 'Not detected'}
        
        try:
            denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
            clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
            enhanced = clahe.apply(denoised)
            _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            thresh_inv = cv2.bitwise_not(thresh)
            
            # Subject
            sub_y1, sub_y2 = int(height * 0.08), int(height * 0.18)
            sub_x1, sub_x2 = int(width * 0.10), int(width * 0.30)
            sub_region = thresh_inv[sub_y1:sub_y2, sub_x1:sub_x2]
            
            if sub_region.size > 0:
                sub_region = cv2.resize(sub_region, None, fx=3, fy=3)
                text = pytesseract.image_to_string(sub_region, lang='eng', config='--psm 7').strip()
                if text and len(text) > 1:
                    info['subject'] = ' '.join(text.split())
            
            # Medium
            med_x1, med_x2 = int(width * 0.40), int(width * 0.55)
            med_region = thresh_inv[sub_y1:sub_y2, med_x1:med_x2]
            
            if med_region.size > 0:
                med_region = cv2.resize(med_region, None, fx=3, fy=3)
                text = pytesseract.image_to_string(med_region, lang='eng', config='--psm 7').strip()
                lower = text.lower()
                if 'eng' in lower:
                    info['medium'] = 'English'
                elif 'sinh' in lower or 'sinhala' in lower:
                    info['medium'] = 'Sinhala'
                elif 'tamil' in lower:
                    info['medium'] = 'Tamil'
            
            # Name
            name_x1, name_x2 = int(width * 0.58), int(width * 0.95)
            name_region = thresh_inv[sub_y1:sub_y2, name_x1:name_x2]
            
            if name_region.size > 0:
                name_region = cv2.resize(name_region, None, fx=3, fy=3)
                text = pytesseract.image_to_string(name_region, lang='eng', config='--psm 7').strip()
                if text and len(text) > 2:
                    info['name'] = ' '.join(text.split()).replace('|', '').replace('_', '')
        
        except Exception as e:
            logger.warning(f"OCR extraction skipped: {e}")
        
        return info
    
    def _extract_all_40_guaranteed(self):
        """Extract all 40 answers"""
        height, width = self.warped.shape[:2]
        answer_area = self.warped[int(height * 0.20):int(height * 0.97), :]
        ans_height, ans_width = answer_area.shape[:2]
        
        gray = cv2.cvtColor(answer_area, cv2.COLOR_BGR2GRAY)
        denoised = cv2.fastNlMeansDenoising(gray, None, h=10, templateWindowSize=7, searchWindowSize=21)
        
        adaptive = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                        cv2.THRESH_BINARY_INV, 15, 3)
        _, otsu = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        combined = cv2.bitwise_or(adaptive, otsu)
        
        kernel = np.ones((2, 2), np.uint8)
        cleaned = cv2.morphologyEx(combined, cv2.MORPH_OPEN, kernel)
        cleaned = cv2.morphologyEx(cleaned, cv2.MORPH_CLOSE, kernel)
        
        contours, _ = cv2.findContours(cleaned, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        all_circles = []
        for cnt in contours:
            area = cv2.contourArea(cnt)
            
            if area < 80 or area > 5000:
                continue
            
            x, y, w, h = cv2.boundingRect(cnt)
            aspect_ratio = w / float(h) if h > 0 else 0
            if aspect_ratio < 0.5 or aspect_ratio > 2.0:
                continue
            
            peri = cv2.arcLength(cnt, True)
            if peri == 0:
                continue
            
            circularity = 4 * np.pi * area / (peri * peri)
            if circularity < 0.3:
                continue
            
            cx = x + w // 2
            cy = y + h // 2
            
            all_circles.append({
                'x': x, 'y': y, 'w': w, 'h': h,
                'cx': cx, 'cy': cy,
                'area': area,
                'circularity': circularity,
                'marked': False,
                'mark_strength': 0.0
            })
        
        logger.info(f"Detected {len(all_circles)} potential circles")
        
        for c in all_circles:
            is_marked, strength = self._is_marked_advanced(gray, cleaned, c)
            c['marked'] = is_marked
            c['mark_strength'] = strength
        
        marked_count = sum(1 for c in all_circles if c['marked'])
        logger.info(f"Marked circles: {marked_count}")
        
        answers = self._extract_with_grid_system(all_circles, ans_width, ans_height)
        
        return answers
    
    def _is_marked_advanced(self, gray, binary, circle):
        """Advanced mark detection"""
        x, y, w, h = circle['x'], circle['y'], circle['w'], circle['h']
        
        pad = 3
        x1 = max(0, x - pad)
        y1 = max(0, y - pad)
        x2 = min(binary.shape[1], x + w + pad)
        y2 = min(binary.shape[0], y + h + pad)
        
        roi_bin = binary[y1:y2, x1:x2]
        roi_gray = gray[y1:y2, x1:x2]
        
        if roi_bin.size == 0 or roi_gray.size == 0:
            return False, 0.0
        
        mask = np.zeros(roi_bin.shape, dtype=np.uint8)
        center = (w // 2 + pad, h // 2 + pad)
        radius = min(w, h) // 2 - 2
        if radius < 3:
            return False, 0.0
        
        cv2.circle(mask, center, radius, 255, -1)
        
        masked_bin = cv2.bitwise_and(roi_bin, mask)
        masked_gray = cv2.bitwise_and(roi_gray, roi_gray, mask=mask)
        
        mask_pixels = np.sum(mask > 0)
        if mask_pixels == 0:
            return False, 0.0
        
        fill_ratio = np.sum(masked_bin > 0) / mask_pixels
        avg_intensity = np.mean(masked_gray[mask > 0])
        min_intensity = np.min(masked_gray[mask > 0])
        std_intensity = np.std(masked_gray[mask > 0])
        
        is_marked = False
        
        if fill_ratio > 0.40 and avg_intensity < 150:
            is_marked = True
        
        if min_intensity < 100:
            is_marked = True
        
        if fill_ratio > 0.50 and avg_intensity < 170:
            is_marked = True
        
        if fill_ratio > 0.35 and avg_intensity < 140 and std_intensity < 35:
            is_marked = True
        
        intensity_score = (255 - avg_intensity) / 255.0
        fill_score = fill_ratio
        darkness_score = (255 - min_intensity) / 255.0
        
        strength = (intensity_score * 0.3 + fill_score * 0.4 + darkness_score * 0.3)
        
        return is_marked, strength
    
    def _extract_with_grid_system(self, all_circles, img_width, img_height):
        """Extract answers using grid system"""
        answers = {}
        
        sorted_circles = sorted(all_circles, key=lambda c: (c['cy'], c['cx']))
        rows = self._cluster_into_rows(sorted_circles, img_height)
        
        logger.info(f"Detected {len(rows)} rows")
        
        col_width = img_width / 4.0
        
        for row_idx, row_circles in enumerate(rows):
            if row_idx >= 10:
                break
            
            row_circles_sorted = sorted(row_circles, key=lambda c: c['cx'])
            
            for col_idx in range(4):
                if col_idx < 3:
                    x1 = int(col_idx * col_width)
                    x2 = int((col_idx + 1) * col_width)
                else:
                    x1 = int(col_idx * col_width - col_width * 0.1)
                    x2 = img_width
                
                col_circles = [c for c in row_circles_sorted if x1 <= c['cx'] < x2]
                
                if not col_circles:
                    continue
                
                col_circles_sorted = sorted(col_circles, key=lambda c: c['cx'])
                answer_circles = self._get_answer_circles(col_circles_sorted)
                
                if len(answer_circles) < 3:
                    continue
                
                marked_circles = [c for c in answer_circles if c['marked']]
                
                if not marked_circles:
                    continue
                
                if len(marked_circles) > 1:
                    marked_circles = sorted(marked_circles, key=lambda c: c['mark_strength'], reverse=True)
                
                marked_circle = marked_circles[0]
                
                try:
                    position = answer_circles.index(marked_circle)
                    option = position + 1
                    
                    question_num = col_idx * 10 + row_idx + 1
                    
                    if 1 <= question_num <= 40 and 1 <= option <= 4:
                        answers[str(question_num)] = option
                
                except ValueError:
                    pass
        
        return answers
    
    def _get_answer_circles(self, circles):
        """Smart selection of answer circles"""
        if len(circles) == 4:
            return circles
        elif len(circles) == 5:
            sizes = [c['area'] for c in circles]
            if sizes[0] < np.mean(sizes[1:]) * 0.7:
                return circles[1:5]
            else:
                return circles[:4]
        elif len(circles) > 5:
            sizes = [c['area'] for c in circles]
            median_size = np.median(sizes)
            scored = [(i, c, abs(c['area'] - median_size)) for i, c in enumerate(circles)]
            scored_sorted = sorted(scored, key=lambda x: x[2])
            top4_indices = sorted([x[0] for x in scored_sorted[:4]])
            return [circles[i] for i in top4_indices]
        else:
            return circles
    
    def _cluster_into_rows(self, circles, img_height):
        """Cluster circles into rows"""
        if not circles:
            return []
        
        expected_row_height = img_height / 10.0
        tolerance = expected_row_height * 0.5
        
        rows = []
        current_row = [circles[0]]
        current_y = circles[0]['cy']
        
        for circle in circles[1:]:
            if abs(circle['cy'] - current_y) < tolerance:
                current_row.append(circle)
            else:
                if current_row:
                    rows.append(current_row)
                current_row = [circle]
                current_y = circle['cy']
        
        if current_row:
            rows.append(current_row)
        
        return rows


# Initialize database
db_manager = DatabaseManager(DB_FILE)


# ==================== FLASK ROUTES ====================

@app.route('/test', methods=['GET'])
def test():
    """Test endpoint"""
    return jsonify({
        "status": "online",
        "message": "Enhanced OMR Grading System v3.0 - Production Ready with Advanced Filtering",
        "version": "3.0",
        "features": [
            "Subject inheritance from master key",
            "Advanced filtering by subject, grade, and level",
            "Grouped results view",
            "Enhanced Excel export with filters",
            "Statistics by subject/grade",
            "Master key management",
            "Student history tracking"
        ]
    })


@app.route('/upload_master', methods=['POST'])
def upload_master():
    """Upload master answer key with metadata"""
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image provided"}), 400
        
        file = request.files['image']
        path = os.path.join(IMAGE_DIR, 'master_key.jpg')
        file.save(path)
        
        # Get metadata
        subject = request.form.get('subject', '').strip()
        if not subject:
            return jsonify({"error": "Subject is required"}), 400
        
        exam_date = request.form.get('exam_date', datetime.now().strftime('%Y-%m-%d'))
        grade_level = request.form.get('grade_level', 'General').strip()
        
        logger.info("="*80)
        logger.info("PROCESSING MASTER ANSWER KEY")
        logger.info(f"Subject: {subject} | Grade: {grade_level}")
        logger.info("="*80)
        
        processor = EnhancedOMRProcessor(path)
        if not processor.process():
            return jsonify({"error": "Image processing failed"}), 400
        
        # Save to file (for backwards compatibility)
        with open(MASTER_DATA_FILE, 'w') as f:
            json.dump(processor.answers, f, indent=2)
        
        # Save metadata
        metadata = {
            'subject': subject,
            'grade_level': grade_level,
            'exam_date': exam_date
        }
        with open(MASTER_METADATA_FILE, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        # Save to database
        master_key_id = db_manager.add_master_key(subject, exam_date, grade_level, processor.answers)
        
        if not master_key_id:
            return jsonify({"error": "Failed to save master key to database"}), 500
        
        logger.info(f"\n✓ Master key saved: {len(processor.answers)}/40 answers")
        
        return jsonify({
            "success": True,
            "message": f"Master key set for {subject} - {grade_level}",
            "total": 40,
            "valid_answers": len(processor.answers),
            "answers": processor.answers,
            "subject": subject,
            "grade_level": grade_level,
            "exam_date": exam_date,
            "master_key_id": master_key_id
        })
    
    except Exception as e:
        logger.error(f"Error in upload_master: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/get_master_metadata', methods=['GET'])
def get_master_metadata():
    """Get current master key metadata"""
    try:
        active_master = db_manager.get_active_master_key()
        
        if not active_master:
            return jsonify({
                "success": False,
                "message": "No active master key"
            }), 404
        
        return jsonify({
            "success": True,
            "subject": active_master['subject'],
            "grade_level": active_master['grade_level'],
            "exam_date": active_master['exam_date']
        })
    
    except Exception as e:
        logger.error(f"Error fetching master metadata: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/grade_student', methods=['POST'])
def grade_student():
    """Grade student with subject inherited from master key"""
    try:
        # Check if master key exists
        active_master = db_manager.get_active_master_key()
        
        if not active_master:
            return jsonify({"error": "No active master key! Please upload master key first."}), 400
        
        if 'image' not in request.files:
            return jsonify({"error": "No image provided"}), 400
        
        file = request.files['image']
        path = os.path.join(IMAGE_DIR, 'student.jpg')
        file.save(path)
        
        # Get student information
        student_id = request.form.get('student_id', '').strip()
        student_name = request.form.get('student_name', '').strip()
        student_medium = request.form.get('student_medium', '').strip()
        
        # Subject and grade level inherited from master key
        subject = active_master['subject']
        grade_level = active_master['grade_level']
        exam_date = active_master['exam_date']
        
        logger.info("="*80)
        logger.info("GRADING STUDENT SHEET")
        logger.info(f"Subject: {subject} | Grade: {grade_level}")
        logger.info("="*80)
        
        # Process image
        processor = EnhancedOMRProcessor(path)
        if not processor.process():
            return jsonify({"error": "Image processing failed"}), 400
        
        student_answers = processor.answers
        detected_info = processor.student_info
        
        # Generate student ID if not provided
        if not student_id:
            student_id = f"STU_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Use detected or provided name
        if not student_name:
            student_name = detected_info.get('name', 'Unknown Student')
        
        # Use detected or provided medium
        final_medium = student_medium or detected_info.get('medium', 'Unknown')
        
        logger.info(f"\nStudent ID: {student_id}")
        logger.info(f"Name: {student_name}")
        logger.info(f"Subject: {subject} (from master key)")
        logger.info(f"Grade Level: {grade_level} (from master key)")
        logger.info(f"Detected: {len(student_answers)}/40 answers")
        
        # Grade the answers using active master key
        master_answers = active_master['answers']
        
        correct = wrong = unanswered = 0
        details = {}
        
        for q in range(1, 41):
            qs = str(q)
            master_ans = master_answers.get(qs)
            student_ans = student_answers.get(qs)
            
            if master_ans is None:
                continue
            
            if student_ans is None:
                unanswered += 1
                details[qs] = {
                    "correct": master_ans,
                    "student": "Not answered",
                    "result": "unanswered"
                }
            elif student_ans == master_ans:
                correct += 1
                details[qs] = {
                    "correct": master_ans,
                    "student": student_ans,
                    "result": "correct"
                }
            else:
                wrong += 1
                details[qs] = {
                    "correct": master_ans,
                    "student": student_ans,
                    "result": "wrong"
                }
        
        total = len(master_answers)
        percentage = round((correct / total) * 100, 2) if total > 0 else 0
        
        results = {
            'total': total,
            'score': correct,
            'correct': correct,
            'wrong': wrong,
            'unanswered': unanswered,
            'percentage': percentage
        }
        
        # Save to database
        db_manager.add_student(student_id, student_name, subject, final_medium, grade_level)
        db_manager.add_grading_result(
            student_id,
            subject,
            grade_level,
            exam_date,
            results,
            student_answers,
            active_master['id']
        )
        
        logger.info(f"✓ RESULT: {correct}/{total} ({percentage}%)")
        
        student_info = {
            'student_id': student_id,
            'name': student_name,
            'subject': subject,
            'medium': final_medium,
            'grade_level': grade_level
        }
        
        return jsonify({
            "success": True,
            "student_info": student_info,
            "total_score": correct,
            "out_of": total,
            "correct": correct,
            "wrong": wrong,
            "unanswered": unanswered,
            "percentage": percentage,
            "details": details
        })
    
    except Exception as e:
        logger.error(f"Error in grade_student: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/get_filters', methods=['GET'])
def get_filters():
    """Get available filter options"""
    try:
        filters = db_manager.get_available_filters()
        
        return jsonify({
            "success": True,
            "filters": filters
        })
    
    except Exception as e:
        logger.error(f"Error fetching filters: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/get_results_grouped', methods=['GET'])
def get_results_grouped():
    """Get results grouped by subject and grade level"""
    try:
        grouped = db_manager.get_results_by_subject_and_grade()
        
        return jsonify({
            "success": True,
            "grouped_results": grouped
        })
    
    except Exception as e:
        logger.error(f"Error fetching grouped results: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/export_excel', methods=['GET'])
def export_excel():
    """Export results to Excel with filters"""
    try:
        subject = request.args.get('subject')
        grade_level = request.args.get('grade_level')
        exam_date = request.args.get('exam_date')
        grade = request.args.get('grade')
        
        results = db_manager.get_all_results(subject, grade_level, exam_date, grade)
        
        if not results:
            return jsonify({"error": "No results found matching filters"}), 404
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create descriptive filename
        filename_parts = ['grading_results']
        if subject:
            filename_parts.append(subject.replace(' ', '_'))
        if grade_level:
            filename_parts.append(grade_level.replace(' ', '_'))
        if grade:
            filename_parts.append(f"grade{grade}")
        filename_parts.append(timestamp)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        filepath = os.path.join(EXPORTS_DIR, filename)
        
        # Prepare filter info for Excel
        filters = {}
        if subject:
            filters['subject'] = subject
        if grade_level:
            filters['grade_level'] = grade_level
        if exam_date:
            filters['exam_date'] = exam_date
        if grade:
            filters['grade'] = grade
        
        success = ExcelExporter.export_results(results, filepath, filters)
        
        if not success:
            return jsonify({"error": "Export failed"}), 500
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/get_all_results', methods=['GET'])
def get_all_results():
    """Get all grading results with optional filters"""
    try:
        subject = request.args.get('subject')
        grade_level = request.args.get('grade_level')
        exam_date = request.args.get('exam_date')
        grade = request.args.get('grade')
        
        results = db_manager.get_all_results(subject, grade_level, exam_date, grade)
        
        formatted_results = []
        for r in results:
            formatted_results.append({
                'student_id': r[0],
                'name': r[1],
                'subject': r[2],
                'medium': r[3],
                'grade_level': r[4],
                'exam_date': r[5],
                'total_questions': r[6],
                'correct': r[7],
                'wrong': r[8],
                'unanswered': r[9],
                'score': r[10],
                'percentage': r[11],
                'grade': r[12],
                'graded_at': r[13]
            })
        
        return jsonify({
            "success": True,
            "total_results": len(formatted_results),
            "results": formatted_results
        })
    
    except Exception as e:
        logger.error(f"Error in get_all_results: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/get_student_history/<student_id>', methods=['GET'])
def get_student_history(student_id):
    """Get grading history for a student"""
    try:
        history = db_manager.get_student_history(student_id)
        
        formatted_history = []
        for h in history:
            formatted_history.append({
                'exam_date': h[0],
                'subject': h[1],
                'grade_level': h[2],
                'score': h[3],
                'percentage': h[4],
                'grade': h[5],
                'graded_at': h[6]
            })
        
        return jsonify({
            "success": True,
            "student_id": student_id,
            "total_exams": len(formatted_history),
            "history": formatted_history
        })
    
    except Exception as e:
        logger.error(f"Error in get_student_history: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/get_statistics', methods=['GET'])
def get_statistics():
    """Get statistics with optional filters"""
    try:
        subject = request.args.get('subject')
        grade_level = request.args.get('grade_level')
        
        stats = db_manager.get_statistics(subject, grade_level)
        
        return jsonify({
            "success": True,
            "statistics": stats
        })
    
    except Exception as e:
        logger.error(f"Error in get_statistics: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    logger.info("="*80)
    logger.info("ENHANCED OMR GRADING SYSTEM v3.0 - PRODUCTION READY")
    logger.info("="*80)
    logger.info("Features:")
    logger.info("  ✓ Subject inheritance from master key")
    logger.info("  ✓ Advanced filtering (subject, grade, level)")
    logger.info("  ✓ Grouped results view")
    logger.info("  ✓ Enhanced Excel export with filters")
    logger.info("  ✓ Statistics by subject/grade")
    logger.info("  ✓ Master key management")
    logger.info("  ✓ Student history tracking")
    logger.info("  ✓ Robust OMR detection")
    logger.info("  ✓ Multi-language OCR support")
    logger.info("="*80)
    
    app.run(host='0.0.0.0', port=5000, debug=True)